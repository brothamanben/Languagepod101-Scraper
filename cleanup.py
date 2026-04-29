import csv
import os
import random
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


SCRIPT_DIR = Path(__file__).resolve().parent
SUPPORTED_EXTENSIONS = {".csv", ".tsv", ".xlsx"}
TYPE_COL = 0
FRONT_COL = 1
BACK_COL = 2
CHOICES_COL = 3
AUDIO_COL = 4
LEVEL_TAG_COL = 5
MIN_COLUMNS = 6


def clean_speaker_labels(text):
    text = re.sub(r"(^|\s)[A-D]:\s*", " ", str(text))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_quotes(text):
    text = str(text)
    return (
        text.replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("â€œ", '"')
        .replace("â€", '"')
        .replace("â€˜", "'")
        .replace("â€™", "'")
    )


def extract_quoted_english(text):
    matches = re.findall(r'"([^"]+)"', normalize_quotes(text))
    return " / ".join(match.strip() for match in matches if match.strip())


def remove_quoted_text(text):
    text = re.sub(r'"[^"]+"', " ", normalize_quotes(text))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def split_into_chunks(text, count=4):
    words = text.split()

    if len(words) >= count:
        base_size, remainder = divmod(len(words), count)
        chunks = []
        start = 0

        for index in range(count):
            current_size = base_size + (1 if index < remainder else 0)
            piece = words[start:start + current_size]
            chunks.append(" ".join(piece))
            start += current_size
    else:
        chars = list(re.sub(r"\s+", "", text))
        if len(chars) >= count:
            base_size, remainder = divmod(len(chars), count)
            chunks = []
            start = 0

            for index in range(count):
                current_size = base_size + (1 if index < remainder else 0)
                piece = chars[start:start + current_size]
                chunks.append("".join(piece))
                start += current_size
        else:
            chunks = chars[:]

    while len(chunks) < count:
        chunks.append("")

    return chunks[:count]


def make_choices(text):
    chunks = [chunk for chunk in split_into_chunks(text, 4) if chunk]
    random.shuffle(chunks)

    while len(chunks) < 4:
        chunks.append("")

    return " ".join(
        f"{chr(65 + i)}) {chunk}".strip()
        for i, chunk in enumerate(chunks)
    ).strip()


def looks_like_audio(text):
    value = str(text).strip()
    lowered = value.lower()
    return (
        lowered.startswith("[sound:")
        or lowered.endswith((".mp3", ".m4a", ".wav", ".ogg"))
    )


def is_header_row(row):
    normalized = [str(cell).strip().lower() for cell in row[:6]]
    return (
        len(normalized) >= 4
        and normalized[0] == "type"
        and normalized[1] == "front"
        and normalized[2] == "back"
        and "audio" in normalized
    )


def has_existing_choices_column(row):
    if len(row) <= CHOICES_COL:
        return False

    cell_value = str(row[CHOICES_COL]).strip().lower()
    return cell_value in {"d", "choices", "choice"}


def has_existing_level_tag_column(row):
    if len(row) <= LEVEL_TAG_COL:
        return False

    cell_value = str(row[LEVEL_TAG_COL]).strip().lower()
    return cell_value in {"leveltag", "level tag", "tag", "tags"}


def extract_level_tag(type_value):
    text = str(type_value).strip()
    if not text:
        return ""

    language_match = re.match(r"\s*([^-]+?)\s*-\s*", text)
    level_match = re.search(r"\bL\s*([0-9]+)\b", text, flags=re.IGNORECASE)

    if not language_match or not level_match:
        return ""

    language = re.sub(r"[^a-z0-9]+", "", language_match.group(1).lower())
    level_number = level_match.group(1)

    if not language or not level_number:
        return ""

    return f"{language}pod101level{level_number}"


def ensure_choices_column(row, header=False):
    row = ["" if value is None else str(value) for value in row]

    while len(row) < 4:
        row.append("")

    if len(row) == 4 and not has_existing_choices_column(row):
        row.insert(CHOICES_COL, "Choices" if header else "")
    else:
        while len(row) < MIN_COLUMNS:
            row.append("")

    if header:
        row[TYPE_COL] = row[TYPE_COL] or "Type"
        row[FRONT_COL] = "Front"
        row[BACK_COL] = "Back"
        row[CHOICES_COL] = "Choices"

        if len(row) <= AUDIO_COL:
            row.append("Audio")
        elif not row[AUDIO_COL]:
            row[AUDIO_COL] = "Audio"
        elif row[AUDIO_COL].strip().lower() != "audio":
            row.insert(AUDIO_COL, "Audio")

        while len(row) <= LEVEL_TAG_COL:
            row.append("")

        if not has_existing_level_tag_column(row):
            row[LEVEL_TAG_COL] = "LevelTag"

        return row

    if looks_like_audio(row[BACK_COL]) and not looks_like_audio(row[AUDIO_COL]):
        row[AUDIO_COL] = row[BACK_COL]
        row[BACK_COL] = ""

    if looks_like_audio(row[CHOICES_COL]) and not looks_like_audio(row[AUDIO_COL]):
        row[AUDIO_COL] = row[CHOICES_COL]
        row[CHOICES_COL] = ""

    return row


def clean_data_row(row):
    row = ensure_choices_column(row, header=False)

    front = str(row[FRONT_COL]).strip()
    back = str(row[BACK_COL]).strip()
    choices = str(row[CHOICES_COL]).strip()

    if not front:
        return row

    front = normalize_quotes(front)
    front = clean_speaker_labels(front)

    extracted_english = extract_quoted_english(front)
    cleaned_front = remove_quoted_text(front)

    if cleaned_front:
        row[FRONT_COL] = cleaned_front

    if extracted_english and (not back or back == front):
        row[BACK_COL] = extracted_english
    elif back:
        row[BACK_COL] = normalize_quotes(back).strip()

    if not choices and cleaned_front:
        row[CHOICES_COL] = make_choices(cleaned_front)

    while len(row) <= LEVEL_TAG_COL:
        row.append("")

    if not row[LEVEL_TAG_COL].strip():
        row[LEVEL_TAG_COL] = extract_level_tag(row[TYPE_COL])

    return row


def sniff_delimiter(path):
    sample = path.read_text(encoding="utf-8-sig", errors="ignore")[:4096]

    if "\t" in sample and sample.count("\t") >= sample.count(","):
        return "\t"

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        return dialect.delimiter
    except csv.Error:
        return ","


def load_delimited_rows(path):
    delimiter = sniff_delimiter(path)

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter))

    return rows, delimiter


def save_delimited_rows(rows, path, delimiter):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)


def unmerge_and_collect_rows(path):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to process .xlsx files.")

    workbook = load_workbook(path)
    sheet = workbook.active
    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_range))

        for row_index in range(min_row, max_row + 1):
            for col_index in range(min_col, max_col + 1):
                sheet.cell(row=row_index, column=col_index).value = top_left_value

    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value) for value in row])

    return workbook, sheet, rows, len(merged_ranges)


def save_xlsx_rows(workbook, sheet, rows, output_path):
    max_existing_rows = max(sheet.max_row, len(rows))
    max_existing_cols = max(sheet.max_column, max((len(row) for row in rows), default=0))

    for row_index in range(1, max_existing_rows + 1):
        for col_index in range(1, max_existing_cols + 1):
            sheet.cell(row=row_index, column=col_index).value = None

    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index).value = value

    workbook.save(output_path)


def process_rows(rows):
    processed = []

    for index, row in enumerate(rows):
        row = ["" if value is None else str(value) for value in row]

        if not any(cell.strip() for cell in row):
            processed.append(row)
            continue

        if index == 0 and is_header_row(row):
            processed.append(ensure_choices_column(row, header=True))
            continue

        processed.append(clean_data_row(row))

    return processed


def process_file(input_path, output_path=None, overwrite=False):
    path = Path(input_path).expanduser().resolve()
    ext = path.suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError("Supported file types are .csv, .tsv, and .xlsx only.")

    if output_path is None:
        if overwrite:
            output_path = path
        else:
            output_path = path.with_name(f"{path.stem}_cleaned{path.suffix}")
    else:
        output_path = Path(output_path).expanduser().resolve()

    if ext == ".xlsx":
        workbook, sheet, rows, merged_count = unmerge_and_collect_rows(path)
        processed_rows = process_rows(rows)
        save_xlsx_rows(workbook, sheet, processed_rows, output_path)
        return output_path, merged_count

    rows, delimiter = load_delimited_rows(path)
    processed_rows = process_rows(rows)
    save_delimited_rows(processed_rows, output_path, delimiter)
    return output_path, 0


def find_candidate_files():
    files = []

    for item in SCRIPT_DIR.iterdir():
        if item.is_file() and item.suffix.lower() in SUPPORTED_EXTENSIONS:
            files.append(item)

    return sorted(files, key=lambda path: path.name.lower())


def resolve_input_path(user_input):
    raw = user_input.strip().strip('"')

    if not raw:
        return None

    direct_path = Path(raw).expanduser()
    if direct_path.exists():
        return direct_path.resolve()

    local_path = SCRIPT_DIR / raw
    if local_path.exists():
        return local_path.resolve()

    candidates = find_candidate_files()
    lowered = raw.lower()

    for candidate in candidates:
        if candidate.name.lower() == lowered:
            return candidate.resolve()

    if "." not in raw:
        for candidate in candidates:
            if candidate.stem.lower() == lowered:
                return candidate.resolve()

    return None


def main():
    candidates = find_candidate_files()

    if candidates:
        print("Files in this folder:")
        for file_path in candidates:
            print("-", file_path.name)
        print()
        prompt = "Type a filename from this folder, or paste a full path: "
    else:
        prompt = "Enter path to your CSV/TSV/XLSX file: "

    user_input = input(prompt)
    input_path = resolve_input_path(user_input)

    if input_path is None:
        print("File not found.")
        return

    try:
        output_path, merged_count = process_file(input_path)
    except Exception as exc:
        print("Error:", exc)
        return

    if merged_count:
        print(f"Unmerged {merged_count} merged range(s).")

    print("Done.")
    print("Saved as:", output_path)


if __name__ == "__main__":
    main()
